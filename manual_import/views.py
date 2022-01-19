import os
import re
import os.path
import shutil
import json
import pandas as pd
import xlsxwriter
from os import listdir
from os.path import isfile, join
import logging

from time import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.workbook.protection import WorkbookProtection

from enum import IntEnum
from datetime import datetime, timedelta

from django.utils import timezone
from django.db import connection
from django.db.utils import Error
from django.conf import settings
from django.http import HttpResponse
from django_filters import rest_framework as filters

from rest_framework import status, generics
from rest_framework.decorators import action, api_view
from rest_framework import viewsets
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.settings import api_settings

from locales.translates_function import translate_label
from SmartEnergy.auth import modules, permissions, check_module
from .models import Ksb1, Fbl34N, UploadFile, GaugeDataTemp
from .serializers import UploadFileSerializer, ListUploadFileSerializer
from gauge_point.models import ImportSource, SourcePme

from .utils.error_manager import ErrorManager
from .utils.verifiers import verify_data_ksb1, verify_data_fbl3n
from .utils.verifiers import verify_file_name, verify_data_ccee
from .utils.verifiers import verify_type_file, verify_data_prod_dash

from rest_framework.pagination import PageNumberPagination

pdct_list = [u'Produção Orçado',
             u'Produção Projetado',
             u'Produção Realizada'
             ]

dshb_list = ['Consumo Orçado (MWm)',
             'Consumo Projetado (MWh)',
             'Custo Energia Orçado (R$)',
             'Custo Energia Real (R$)',
             'Demanda Contratada Ponta (kW)',
             'Demanda Contratada Fora Ponta (kW)',
             'Demanda Orçada Ponta (kW)',
             'Demanda Orçada Fora Ponta (kW)',
             'Gasto Demanda Orçado (R$)',
             'Gasto Demanda Real (R$)',
             'Gasto Encargos Orçado (R$)',
             'Gasto Encargos Real (R$)'
            ]


class ManualImport(IntEnum):
    PRODUCAO = 0
    DASHBOARD = 1
    CCEE = 2
    KSB1 = 3
    FBL34N = 4
    CCEE_CSV = 5


class Status(IntEnum):
    OK = 1
    ONE_ERROR = 0
    MULTI_ERRORS = 2
    IN_PROGRESS = 3


class CustomPaginationClass(PageNumberPagination):
    page_size_query_param = 'page_size'


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class UploadFileFilter(filters.FilterSet):
    # file_name = filters.CharFilter(lookup_expr='icontains')
    # file_type = filters.CharFilter(lookup_expr='icontains')
    date_upload = filters.CharFilter(lookup_expr='icontains')
    user_name = filters.CharFilter(lookup_expr='icontains')
    send_status = filters.CharFilter(lookup_expr='icontains')
    # msg = filters.CharFilter(lookup_expr='icontains')

    serializer_class = UploadFileSerializer
    pagination_class = StandardResultsSetPagination

    class Meta:
        model = UploadFile
        # fields = ('file_name', 'file_type', 'date_upload', 'user_name', 'send_status', 'msg')
        fields = ('date_upload', 'user_name', 'send_status')


class UploadFileView(generics.ListAPIView):

    queryset = UploadFile.objects.all()
    serializer_class = UploadFileSerializer
    pagination_class = StandardResultsSetPagination


class UploadFileViewSet(viewsets.ModelViewSet):

    parser_classes = (MultiPartParser, FormParser)
    queryset = UploadFile.objects.all()
    serializer_class = UploadFileSerializer
    pagination_class = StandardResultsSetPagination
    filterset_class = UploadFileFilter

    @property
    def get_queryset(self):

        _type = self.request.query_params.get('type', None)
        date_upload = self.request.query_params.get('date_upload', None)
        user_name = self.request.query_params.get('user_name', None)
        send_status = self.request.query_params.get('send_status', None)
        offset = self.request.query_params.get('offset', None)

        if self.request.query_params.get('ordering'):
            ordering = self.request.query_params.get('ordering')
        else:
            ordering = '-date_upload'

        filter_params = {}
        if _type:
            filter_params = {'file_type':_type}

        if date_upload:
            try:
                has_match = False
                hour_tz = int(offset) / 60

                match = re.search("^([0-9][0-9])[-](0[1-9]|1[012])[-](0[1-9]|[12][0-9]|3[01])$", date_upload)
                if match:
                    date_upload = '20'+date_upload

                # Verifica o padrao yyyy-mm-dd
                match = re.search("^(19[0-9][0-9]|20[0-9][0-9])[-](0[1-9]|1[012])[-](0[1-9]|[12][0-9]|3[01])$", date_upload)
                if match:
                    has_match = True

                    start_date_str = date_upload
                    _start_date = datetime.strptime(start_date_str, '%Y-%m-%d') + timedelta(hours=hour_tz)

                    end_date_str = date_upload + ' 23:59:59'
                    _end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S') + timedelta(hours=hour_tz)
                    filter_params['date_upload__range'] = (_start_date, _end_date)

                # Verifica o padrao -mm-dd
                match = re.search("^[-]?(0[1-9]|1[012])[-](0[1-9]|[12][0-9]|3[01])$", date_upload)
                if match:
                    has_match = True
                    current_year = str(timezone.now().year)

                    start_date_str = current_year + date_upload
                    _start_date = datetime.strptime(start_date_str, '%Y-%m-%d') + timedelta(hours=hour_tz)

                    end_date_str = current_year + date_upload + ' 23:59:59'
                    _end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S') + timedelta(hours=hour_tz)

                    filter_params['date_upload__range'] = (_start_date, _end_date)

                # Verifica o padrao -dd
                match = re.search("^[-][-]?(0[1-9]|[12][0-9]|3[01])$", date_upload)
                if match:
                    has_match = True
                    _date = datetime.strptime(date_upload, '--%d')
                    filter_params['date_upload__day'] = _date.day
                if not has_match:
                    _date = datetime.strptime("1900-01-01", '%Y-%m-%d')
                    filter_params['date_upload__day'] = _date.day
                    filter_params['date_upload__month'] = _date.month
                    filter_params['date_upload__year'] = _date.year

            except Exception as e:
                print(e.args)
                _date = datetime.strptime("1900-01-01", '%Y-%m-%d')
                filter_params['date_upload__day'] = _date.day
                filter_params['date_upload__month'] = _date.month
                filter_params['date_upload__year'] = _date.year

        if user_name:
            filter_params['user_name__icontains'] = user_name
        if send_status:
            filter_params['send_status__icontains'] = send_status

        queryset = UploadFile.objects.filter(**filter_params).order_by(ordering)
        return queryset

    @check_module(modules.manual_import, [permissions.VIEW, permissions.EDITN1])
    def list(self, request, *args, **kwargs):
        # queryset = self.filter_queryset(self.get_queryset)
        queryset = self.get_queryset
        try:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = ListUploadFileSerializer(page, many=True)
                return self.get_paginated_response(serializer.data)
        except Exception as e:
            print(e.args)
            queryset = self.filter_queryset(self.get_queryset)
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

    def get_serializer(self, *args, **kwargs):
        serializer_class = self.get_serializer_class()
        kwargs['context'] = self.get_serializer_context()
        return serializer_class(*args, **kwargs)

    @check_module(modules.manual_import, [permissions.EDITN1, permissions.VIEW])
    def create(self, request, *args, **kwargs):

        _utc_create = datetime.now(tz=timezone.utc).replace(microsecond=0)
        _user_name = "import-manual"
        _short_name = "import-manual"

        if request.auth:
            _short_name = request.auth['cn']
            _user_name = request.auth['cn'] + ' - ' + request.auth['UserFullName']

        _file_name = str(request.data['file_name'])
        _file_type = str(request.data['file_type'])
        _file_find_type = _file_type
        if 'PDCT' in _file_type or 'producao' in _file_type:
            _file_type = 'producao'
            _file_find_type = 'PDCT'
            request.data['file_type'] = 'PDCT'

        if 'DSHB' in _file_type or 'dashboard' in _file_type:
            _file_type = 'dashboard'
            _file_find_type = 'DSHB'
            request.data['file_type'] = 'DSHB'

        # Pasta de origem de todos os arquivos: /uploads/need_process/
        origin = settings.MEDIA_ROOT + '/need_process/'
        _file_path = origin + _file_name

        # Verificando o template do arquivo com a pagina aberta
        validate_file_type = verify_type_file(_file_path, _file_type)
        if not validate_file_type:

            resp = {'msg': translate_label('error_ima_file_type', request) + _file_type.upper()}
            return Response(resp)

        # Codigo nativo da biblioteca
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        _datetime = str(datetime.now().timestamp()).replace('.', '')
        _new_name = _datetime + "_" + _short_name + "_" + serializer.validated_data['file_name'].split('_')[0]
        _new_name += '.' + serializer.validated_data['file_name'].split(".")[-1]

        serializer.validated_data['file_path'].name = _new_name
        serializer.validated_data['user_name'] = _user_name

        # Salva o regitro e o arquivo na pasta /uploads/need_process/
        self.perform_create(serializer)

        # Verifica se na tabela ImportSource existe o tipo Manual (Importe Manual)
        _ima = ImportSource.objects.filter(import_type='Manual').first()
        if not _ima:
            # Envia mensagem para o front-end
            headers = self.get_success_headers(serializer.data)
            resp = {'msg': translate_label('error_ima_database_IMA', request), 'errors': []}
            return Response(resp, status=status.HTTP_201_CREATED, headers=headers)

        # Verifica o padrao de nomes do arquivo
        folder = verify_file_name(_file_path)
        if folder:

            # Se o arquivo estiver no padrao de nomes correto, entao verifica qual eh sua pasta de destino
            destiny = settings.MEDIA_ROOT + '/' + folder
            _file_path = destiny + _new_name
            origin = settings.MEDIA_ROOT + '/need_process/' + _new_name

            # Pega o ultimo registro inserido pelo usuario _user_name
            try:
                upload = UploadFile.objects.get(file_path='need_process/' + _new_name,
                                                user_name=_user_name,
                                                file_type=_file_find_type)

                upload.send_status = Status.IN_PROGRESS

                # Se o arquivo jah existir na pasta entao delete
                if os.path.isfile(_file_path):
                    os.remove(_file_path)

                # Se o diretorio nao exsitir no servidor, entao sera criado
                if not os.path.exists(destiny):
                    os.makedirs(destiny)

                # Move o novo arquivo para a pasta certa
                shutil.move(origin, destiny)

                # Atualiza o objeto com o novo caminho para a pasta
                upload.date_upload = _utc_create
                upload.save()

                csv_file = False
                df = None
                error_manager = ErrorManager()
                try:
                    f_name, f_extension = os.path.splitext(_file_path)
                    if '_ccee' in f_name.lower() and '.csv' in f_extension:
                        csv_file = True

                    if csv_file:
                        df = pd.read_csv(_file_path,
                                         encoding="ISO-8859-1",
                                         error_bad_lines=False,
                                         skiprows=3,
                                         sep=';',
                                         usecols=[1, 2, 3, 4, 6]
                                         )
                    else:
                        df = pd.read_excel(_file_path, 'Dados')

                except Exception as e:
                    print(e.args)
                    if csv_file:
                        error_manager.set_single_error("error_ima_csv_wrong")
                    else:
                        error_manager.set_single_error("error_ima_template_tab")

                if not error_manager.has_single_errors():
                    # Valida o conteudo do arquivo
                    if 'ksb1' in folder:
                        error_manager = validate_file(df, ManualImport.KSB1, upload)
                    if 'fbl3' in folder:
                        error_manager = validate_file(df, ManualImport.FBL34N)
                    if 'ccee' in folder:
                        if csv_file:
                            error_manager = validate_file_gauge(df, _file_path, _utc_create, ManualImport.CCEE_CSV)
                        else:
                            error_manager = validate_file_gauge(df, _file_path, _utc_create, ManualImport.CCEE)
                    if 'producao' in folder:
                        error_manager = validate_file_gauge(df, _file_path, _utc_create, ManualImport.PRODUCAO)
                    if 'dashboard' in folder:
                        error_manager = validate_file_gauge(df, _file_path, _utc_create, ManualImport.DASHBOARD)

                # Atualiza a mensagem de sucesso ou de erro do objeto
                upload.msg = json.dumps(error_manager.get_dic_response())
                upload.send_status = Status.OK

                if error_manager.has_multi_errors():
                    upload.send_status = Status.MULTI_ERRORS

                if error_manager.has_single_errors():
                    upload.send_status = Status.ONE_ERROR

                upload.save()

                # Deleta o arquivo no final do processamento
                if os.path.isfile(_file_path):
                    os.remove(_file_path)

                # Envia mensagem para o front-end
                headers = self.get_success_headers(serializer.data)
                resp = format_dic_ima_erro(error_manager.get_dic_response(), request)
                return Response(resp, status=status.HTTP_201_CREATED, headers=headers)

            except UploadFile.DoesNotExist:
                # Envia mensagem para o front-end
                resp = {'msg': translate_label('error_ima_file_find', request), 'errors': []}
                return Response(resp, status=status.HTTP_404_NOT_FOUND)
        else:
            # Se o arquivo estiver fora do padrao de nomes, delete
            if os.path.isfile(_file_path):
                os.remove(_file_path)

            # Envia mensagem para o front-end
            headers = self.get_success_headers(serializer.data)
            resp = {'msg': translate_label('error_ima_file_name', request), 'errors': []}
            return Response(resp, status=status.HTTP_201_CREATED, headers=headers)

    @action(methods=['get'], detail=True)
    def get_file_upload(self, request, pk=None, *args, **kwargs):

        _file_type = request.query_params['page']
        _username = request.user.username
        try:
            _file_name = ''
            file_path = settings.MEDIA_ROOT + '/templates/'

            __templates = os.listdir(file_path)
            if __templates is not None and int(pk) == 0:

                for templates in __templates:
                    if _file_type in templates:
                        _file_name = templates
                        break

                full_path = file_path + _file_name
                if '_PDCT' in _file_name or '_DSHB' in _file_name:
                    full_path = load_sources(file_path, _file_name, _username)

                if _file_name.__len__() == 0 or not os.path.exists(file_path):
                    # Retonar vazio porque o template requisitado não existe
                    resp = {'msg': translate_label('error_ima_template', request)}
                    return HttpResponse(resp, content_type="text/plain")

                with open(full_path, 'rb') as file:
                    _file = file.read()

                response = HttpResponse(_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={_file_name}'
                return response
            else:
                # Retonar vazio porque o template requisitado não existe
                resp = {'msg': translate_label('error_ima_template', request)}
                return HttpResponse(resp, content_type="text/plain")

        except Exception as e:
            print(e.args)
            # Erro ao tentar baixar template
            resp = {'msg': translate_label("error_ima_template_download", request)}
            return HttpResponse(resp, content_type="text/plain")

    @check_module(modules.manual_import, [permissions.EDITN1, permissions.VIEW])
    def update(self, request, *args, **kwargs):

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    def perform_update(self, serializer):
        serializer.save()

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    def get_success_headers(self, data):
        try:
            return {'Location': str(data[api_settings.URL_FIELD_NAME])}
        except (TypeError, KeyError):
            return {}


def load_sources(path, file_name, _username):

    # ---------------------------------------------------
    # Get type
    # ---------------------------------------------------
    # New file name
    _type_file = '_DSHB_'
    name_dshb = _username + '_DSHB_'
    name_pdct = _username + '_PDCT_'
    name_file = _username + '_DSHB_'

    if '_PDCT' in file_name:
        _type_file = '_PDCT_'
        name_file = _username + _type_file
    # ---------------------------------------------------
    # Remove old files
    # ---------------------------------------------------

    _files = [f for f in listdir(path) if isfile(join(path, f))]
    for f in _files:
        if os.path.isfile(path+f):
            if name_pdct in f:
                os.remove(path+f)
            if name_dshb in f:
                os.remove(path+f)

    # Set sheet protection key for worksheets
    _sheet_protection_key = 'smartenergy@vale'

    # ---------------------------------------------------
    # Sheet names
    # ---------------------------------------------------
    _sheet_dados_name = 'Dados'
    _sheet_info_name = u'Instruções'
    _sheet_source_name = 'Sources'

    # New file name
    new_name = name_file + str(time()).split('.')[0] + '.xlsx'
    full_path = path+new_name

    # List of sources
    source_list = list(SourcePme.objects.values_list('display_name', flat=True).distinct())

    # Create temporary excel file
    wb = xlsxwriter.Workbook(full_path)

    # Create sheets
    ws_dados = wb.add_worksheet(_sheet_dados_name)
    ws_info = wb.add_worksheet(_sheet_info_name)

    _type = ''
    _image_path = path + 'info_dshb.png'
    quantity_list = dshb_list

    if 'PDCT' in full_path:
        _type = ' (kt)'
        _image_path = path + 'info_pdct.png'
        quantity_list = pdct_list

    # Protect and activated sheet info
    # ws_info.insert_image('A1', _image_path)
    ws_info.hide_gridlines(2)
    ws_info.activate()
    ws_info.protect(_sheet_protection_key)

    # ---------------------------------------------------
    # Sheet source
    # ---------------------------------------------------
    # Load data to sheet source and then hide and protected the sheet

    i = 1
    for item in source_list:
        ws_dados.write_string('X'+str(i), item)
        i += 1

    i = 1
    for item in quantity_list:
        ws_dados.write_string('Z'+str(i), item)
        i += 1

    # ---------------------------------------------------
    # Sheet dados
    # ---------------------------------------------------
    header_format = wb.add_format()
    header_format.set_bold()
    header_format.set_locked()
    header_format.set_align('center')
    header_format.set_valign('center')

    _column_A = u'TAG do ponto de medição'
    _column_B = 'Grandeza'
    _column_C = 'Valor' + _type
    _column_D = 'Data'

    # Set header
    ws_dados.set_column('A1:A1', 50, header_format)
    ws_dados.set_column('B1:B1', 30, header_format)
    ws_dados.set_column('C1:C1', 15, header_format)
    ws_dados.set_column('D1:D1', 15, header_format)

    ws_dados.write('A1', _column_A, header_format)
    ws_dados.write('B1', _column_B, header_format)
    ws_dados.write('C1', _column_C, header_format)
    ws_dados.write('D1', _column_D, header_format)

    # Unlock cells to receive data
    unlocked = wb.add_format({'locked': False})
    ws_dados.set_column('A2:A1048576', 50, unlocked)
    ws_dados.set_column('B2:B1048576', 30, unlocked)
    ws_dados.set_column('C2:C1048576', 15, unlocked)
    ws_dados.set_column('D2:D1048576', 15, unlocked)

    # Create data validation
    ws_dados.set_header()
    _source_data = '=$X$1:$X$'+str(len(source_list))
    _quantity_data = '=$Z$1:$Z$'+str(len(quantity_list))

    ws_dados.data_validation('A2:A1048576', {'validate': 'list', 'source': _source_data})
    ws_dados.data_validation('B2:B1048576', {'validate': 'list', 'source': _quantity_data})
    ws_dados.data_validation('C2:C1048576', {'validate': 'decimal',
                                             'criteria': 'between',
                                             'minimum': 0.0,
                                             'maximum': 9999999999.00,
                                             'error_title': u'Valor inserido não é válido!',
                                             'error_message': u'Valor deve ser decimal com duas casas.'})

    ws_dados.data_validation('D2:D1048576', {'validate': 'date',
                                             'criteria': 'between',
                                             'minimum': date(1900, 1, 1),
                                             'maximum': date(2099, 12, 12),
                                             'error_title': u'Data inválida!',
                                             'error_message': u'A data deve ter o formato dd/mm/yyyy'})

    ws_dados.set_column('E:XFD', None, None, {'hidden': True})
    ws_dados.protect(_sheet_protection_key)
    wb.close()

    # return full_path
    # Insert info image on file and lock the structure of the excel file
    _wb = load_workbook(filename=full_path)
    img = Image(_image_path)

    _ws_info = _wb[_sheet_info_name]
    _ws_info.add_image(img, 'A1')
    _ws_info.protection.set_password(_sheet_protection_key)

    _ws_dados = _wb[_sheet_dados_name]
    _ws_dados.protection.set_password(_sheet_protection_key)

    _wb.security = WorkbookProtection(workbookPassword=_sheet_protection_key, lockStructure=True)
    _wb.save(filename=full_path)

    return full_path


def validate_file(df, _type, upload_file: UploadFile = None):
    
    ksb1 = None
    fbl34n = None

    error_manager = ErrorManager()
    if _type == ManualImport.KSB1:
        list_data, error_manager = verify_data_ksb1(df, error_manager)
    else:
        list_data, error_manager = verify_data_fbl3n(df, error_manager)

    if error_manager.has_single_errors():
        return error_manager

    try:
        for d in list_data:

            if _type is ManualImport.KSB1:
                ksb1 = Ksb1()
                for k, v in d.items():
                    if v is 'NULL':
                        continue
                    setattr(ksb1, k, v)
            else:
                fbl34n = Fbl34N()
                for k, v in d.items():
                    if v is 'NULL':
                        continue
                    setattr(fbl34n, k, v)

            _ima = ImportSource.objects.filter(import_type='Manual')
            utc = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

            if _type is ManualImport.KSB1:
                ksb1.id_import_source = _ima[0].id_import_source
                ksb1.utc_creation = utc
                ksb1.upload_file = upload_file
                ksb1.save()
            else:
                fbl34n.id_import_source = _ima[0].id_import_source
                fbl34n.utc_creation = utc
                fbl34n.save()

        error_manager.set_dic_response()
        return error_manager

    except ValueError as v:
        # Erro ao inserir as informações no banco de dados
        print(v.args)
        error_manager.set_single_error("error_ima_database")
        return error_manager


def validate_file_gauge(df, file_path, _utc, _type):

    list_data = []
    error_manager = ErrorManager()

    if _type == ManualImport.PRODUCAO or _type == ManualImport.DASHBOARD:
        list_data, error_manager = verify_data_prod_dash(df, _utc, file_path, error_manager)
    if _type == ManualImport.CCEE:
        list_data, error_manager = verify_data_ccee(df, _utc, error_manager)
    if _type == ManualImport.CCEE_CSV:
        list_data, error_manager = verify_data_ccee(df, _utc, error_manager, True)

    if error_manager.has_single_errors():
        return error_manager

    if len(list_data) == 0:
        error_manager.set_single_error("error_ima_empty_data")

    try:
        if len(list_data) > 0:
            GaugeDataTemp.objects.bulk_create(list_data)
            error_manager = call_proc(_utc, error_manager)

        if not error_manager.has_single_errors():
            error_manager.set_dic_response()

        return error_manager

    except Exception as ex:
        # Erro ao inserir as informações no banco de dados
        print(ex.args)
        error_manager.set_single_error("error_ima_database")
        return error_manager


def call_proc(_utc, em):

    cursor = connection.cursor()
    _utc_str = _utc.strftime("%Y-%m-%d %H:%M:%S")

    # EXEC[dbo].[spMergeImportGaugeData] @UTC_Creation = '2020-06-30 20:19:01.557'
    _spCommand = "EXEC [dbo].[spMergeImportGaugeData] @UTC_Creation = '" + _utc_str + "'"

    try:
        cursor.execute(_spCommand)
    except Error as e:
        print(e.args)
        logging.error("Error while call the stored procedure spMergeImportGaugeData", exc_info=e)
        
        GaugeDataTemp.objects.filter(utc_creation=_utc).delete()
        em.set_single_error("error_ima_proc")
    finally:
        cursor.close()
        return em


@api_view(['GET'])
def get_ima_erro_log(request, pk):

    dic_response = {}
    _uppload_file = UploadFile.objects.filter(id=pk).first()

    if _uppload_file:
        _msg_json = json.loads(json.dumps(_uppload_file.msg))
        _msg_dic = json.loads(_msg_json)

        # Pegar este formato :
        # "msg": "success_ima_errors",
        # "errors": [
        #      {"error_type": "error_ima_source_gauge", "error_param": "BR_BARBACENA.FRONTEIRA", "lines": '1,2,3,4'},
        #      {"error_type": "error_ima_source_gauge", "error_param": "BR_ARAXA.FRONTEIRA", "lines": '11,12'},
        #      {"error_type": "error_ima_required", "error_param": "Data", "lines": '17,18'}
        # ],

        # E traduzir em uma mensagem neste formato :
        # "msg": "Importado com sucesso. Algumas linhas contém erros.",
        # "errors": ["Linhas '1,2,3,4' - O medidor a seguir não possui ponto de medição cadastrado: BR_BARBACENA.FRONTEIRA",
        #            "Linhas '11,12' - O medidor a seguir não possui ponto de medição cadastrado: BR_ARAXA.FRONTEIRA",
        #            "Linhas '17,18' - Arquivo faltando dados obrigatórios na coluna: Data"]

        _erro_list = []
        if 'errors' in _msg_dic:
            for e in _msg_dic['errors']:
                _erro_aux = translate_label("error_ima_line", request) + " "
                _erro_aux += e['lines'] + " - "
                _erro_aux += translate_label(e['error_type'], request) + " "
                _erro_aux += e['error_param']
                _erro_list.append(_erro_aux)

        dic_response["msg"] = translate_label(_msg_dic["msg"], request)
        dic_response["errors"] = _erro_list
    return Response(dic_response)


def format_dic_ima_erro(dic, request):

    _msg_dic = dic
    dic_saida = {}

    _erro_list = []
    if 'errors' in _msg_dic:
        for e in _msg_dic['errors']:
            _erro_aux = translate_label("error_ima_line", request) + " "
            _erro_aux += e['lines'] + " - "
            _erro_aux += translate_label(e['error_type'], request) + " "
            _erro_aux += e['error_param']
            _erro_list.append(_erro_aux)

    dic_saida["msg"] = translate_label(_msg_dic["msg"], request)
    dic_saida["errors"] = _erro_list
    return dic_saida
