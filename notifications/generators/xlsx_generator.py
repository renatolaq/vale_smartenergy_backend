from io import BytesIO

from django.utils.translation import gettext as _, activate, deactivate
from openpyxl.styles import Font, PatternFill
import openpyxl

from SmartEnergy.settings import BASE_DIR


def generate_xls(data, display_title, language="pt_BR"):
    workbook: openpyxl.Workbook = openpyxl.load_workbook(
        filename="notifications/templates/notification_list_template.xlsx"
    )
    activate(language)
    logo_vale_folder = f"{BASE_DIR}/SmartEnergy/static/vale-logo-xlsx.png"

    default_sheet = workbook.active
    sheet_title = display_title
    logo_vale = openpyxl.drawing.image.Image(logo_vale_folder)

    current_sheet = workbook.copy_worksheet(default_sheet)
    current_sheet.title = _(sheet_title)
    current_sheet.add_image(logo_vale)

    current_row = 2
    current_column = 1
    for field in data.get("fields"):
        cell = current_sheet.cell(
            row=current_row, column=current_column, value=_(field)
        )
        set_header_style(cell)
        current_column += 1

    for notification in data.get("notifications"):
        current_row += 1
        current_column = 1
        for key, value in notification.items():
            cell = current_sheet.cell(
                row=current_row, column=current_column, value=_(value)
            )
            set_content_style(cell)
            current_column += 1

    adjust_column_cell(current_sheet)
    workbook.remove_sheet(default_sheet)

    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)

    workbook_bytes.seek(0)
    excelFile = workbook_bytes.read()
    deactivate()
    return excelFile


def set_content_style(cell):
    font = Font(name="Calibri", size=11)
    cell.font = font


def set_header_style(cell):
    font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="327E7A", end_color="327E7A", fill_type="solid")

    cell.font = font
    cell.fill = fill


def adjust_column_cell(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value * 1.5
