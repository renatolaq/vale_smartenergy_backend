from decimal import Decimal
from typing import List
from copy import deepcopy

from calendar import month_name, different_locale
from io import BytesIO

from ..models.CompanyPlanMonitoring import CompanyPlanMonitoring
from ..models.CompanyPlanMonitoringRevision import CompanyPlanMonitoringRevision

from company.models import Company

import openpyxl


class ReportService:
    def generate_excel_report(self, company_plan_monitorings: List[CompanyPlanMonitoring], language) -> BytesIO:
        company_name_input_col = 1
        company_name_input_row = 2
        year_input_col = 1
        year_input_row = 3
        data_input_start_col = 2
        data_input_start_row = 8
        data_output_start_col = 2
        data_output_start_row = 30
        years_names = ['firstyear_plan_monitoring', 'secondyear_plan_monitoring']
        english_month_names = []
        with different_locale('C.UTF-8'):
            english_month_names = list(map(lambda s: s.lower(), month_name))
            english_month_names.pop(0)

        workbook: openpyxl.Workbook = openpyxl.load_workbook(
            filename=f"plan_monitoring/services/report_templates/plan_monitoring_report_template_{language}.xlsx")

        worksheets: List[openpyxl.Worksheet] = []
        worksheets.append(workbook.active)
        for company_plan_monitoring in company_plan_monitorings[1:]:
            worksheets.append(workbook.copy_worksheet(worksheets[0]))
            worksheets[-1]._images = deepcopy(worksheets[0]._images)

        i = 0
        for company_plan_monitoring in company_plan_monitorings:
            last_revision: CompanyPlanMonitoringRevision = company_plan_monitoring["companyplanmonitoringrevision_set"][0]

            company_name = Company.objects.get(
                id_company=company_plan_monitoring["company_id"]).company_name
            company_plan_monitoring_year = company_plan_monitoring["year"]

            worksheet = worksheets[i]
            worksheet.title = f"{company_plan_monitoring_year}_{company_name}"[:31]
            i += 1

            worksheet.cell(row=company_name_input_row,
                           column=company_name_input_col, value=company_name)
            worksheet.cell(
                row=year_input_row, column=year_input_col).value = \
                (worksheet.cell(
                    row=year_input_row, column=year_input_col).value or "") + \
                f"{company_plan_monitoring_year}"

            def getValue(year, month, field):
                return "-" if getattr(getattr(getattr(last_revision, years_names[year]), english_month_names[month]), field) is None else \
                    getattr(getattr(
                        getattr(last_revision, years_names[year]), english_month_names[month]), field)

            for year in range(1):
                for month in range(12):
                    worksheet.cell(row=data_input_start_row - 1, column=data_input_start_col +
                                   year * 12 + month).value += str(company_plan_monitoring_year + year)
                    worksheet.cell(row=data_input_start_row, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "contracted_peak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 1, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "contracted_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 2, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_peakpower_demand"))
                    worksheet.cell(row=data_input_start_row + 3, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "estimated_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 4, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "projected_peakpower_demand"))
                    worksheet.cell(row=data_input_start_row + 5, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "projected_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 6, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_peakpower_demand"))
                    worksheet.cell(row=data_input_start_row + 7, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 8, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_peak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 9, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_offpeak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 10, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "projected_peak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 11, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "projected_offpeak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 12, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_peak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 13, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_offpeak_power_consumption"))
                    worksheet.cell(row=data_input_start_row + 14, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_production"))
                    worksheet.cell(row=data_input_start_row + 15, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "projected_production"))
                    worksheet.cell(row=data_input_start_row + 16, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_production"))
                    worksheet.cell(row=data_input_start_row + 17, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_productive_stops"))
                    worksheet.cell(row=data_input_start_row + 18, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "projected_productive_stops"))

                    worksheet.cell(row=data_output_start_row - 1, column=data_output_start_col +
                                   year * 12 + month).value += str(company_plan_monitoring_year+year)
                    worksheet.cell(row=data_output_start_row, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_total_consumption"))
                    worksheet.cell(row=data_output_start_row + 1, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "projected_total_consumption"))
                    worksheet.cell(row=data_output_start_row + 2, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "realized_total_consumption"))
                    worksheet.cell(row=data_output_start_row + 3, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "variation_consumption_estimated_realized"))
                    worksheet.cell(row=data_output_start_row + 4, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "variation_consumption_estimated_projected"))
                    worksheet.cell(row=data_output_start_row + 5, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_utilization_factor_consistency_peakpower"))
                    worksheet.cell(row=data_output_start_row + 6, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_utilization_factor_consistency_offpeakpower"))
                    worksheet.cell(row=data_output_start_row + 7, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_load_factor_consistency_peakpower"))
                    worksheet.cell(row=data_output_start_row + 8, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_load_factor_consistency_offpeakpower"))
                    worksheet.cell(row=data_output_start_row + 9, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_uniqueload_factor_consistency"))
                    worksheet.cell(row=data_output_start_row + 10, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_modulation_factor_consistency"))
                    worksheet.cell(row=data_output_start_row + 11, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_specific_consumption"))
                    worksheet.cell(row=data_output_start_row + 12, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "projected_specific_consumption"))
                    worksheet.cell(row=data_output_start_row + 13, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "realized_specific_consumption"))
                    worksheet.cell(row=data_output_start_row + 14, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "variation_specific_consumption_estimated_realized"))
                    worksheet.cell(row=data_output_start_row + 15, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "variation_specific_consumption_estimated_projected"))

        workbook_bytes = BytesIO()
        workbook.save(workbook_bytes)
        return workbook_bytes
