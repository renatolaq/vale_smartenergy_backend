from decimal import Decimal
from typing import List
from copy import deepcopy

from calendar import month_name, different_locale
from io import BytesIO

from ..models.CompanyBudget import CompanyBudget
from ..models.CompanyBudgetRevision import CompanyBudgetRevision

from .IntegrationService import IntegrationService
from company.models import Company

import openpyxl


class ReportService:
    def generate_excel_report(self, company_budgets: List[CompanyBudget], language) -> BytesIO:
        company_name_input_col = 1
        company_name_input_row = 2
        year_input_col = 1
        year_input_row = 3
        data_input_start_col = 2
        data_input_start_row = 8
        data_output_start_col = 2
        data_output_start_row = 19
        years_names = ['firstyear_budget', 'secondyear_budget',
                       'thirdyear_budget', 'fourthyear_budget', 'fifthyear_budget']
        english_month_names = []
        with different_locale('C.UTF-8'):
            english_month_names = list(map(lambda s: s.lower(), month_name))
            english_month_names.pop(0)

        workbook: openpyxl.Workbook = openpyxl.load_workbook(
            filename=f"budget/services/report_templates/budget_report_template_{language}.xlsx")

        worksheets: List[openpyxl.Worksheet] = []
        worksheets.append(workbook.active)
        for company_budget in company_budgets[1:]:
            worksheets.append(workbook.copy_worksheet(worksheets[0]))
            worksheets[-1]._images = deepcopy(worksheets[0]._images)

        i = 0
        for company_budget in company_budgets:
            last_revision: CompanyBudgetRevision = company_budget["companybudgetrevision_set"][0]

            company_name = Company.objects.get(
                id_company=company_budget["company_id"]).company_name
            company_budget_year = company_budget["year"]

            worksheet = worksheets[i]
            worksheet.title = f"{company_budget_year}_{company_name}"[:31]
            i += 1

            worksheet.cell(row=company_name_input_row,
                           column=company_name_input_col, value=company_name)
            worksheet.cell(
                row=year_input_row, column=year_input_col).value += f"{company_budget_year}-{company_budget_year+4}"

            def getValue(year, month, field):
                return "-" if getattr(getattr(getattr(last_revision, years_names[year]), english_month_names[month]), field) is None else \
                    getattr(getattr(
                        getattr(last_revision, years_names[year]), english_month_names[month]), field)

            for year in range(5):
                for month in range(12):
                    worksheet.cell(row=data_input_start_row - 1, column=data_input_start_col +
                                   year * 12 + month).value += str(company_budget_year + year)
                    worksheet.cell(row=data_input_start_row, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "contracted_peak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 1, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "contracted_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 2, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "estimated_peak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 3, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "estimated_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 4, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "consumption_peak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 5, column=data_input_start_col + year *
                                   12 + month, value=getValue(year, month, "consumption_offpeak_power_demand"))
                    worksheet.cell(row=data_input_start_row + 6, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "production"))
                    worksheet.cell(row=data_input_start_row + 7, column=data_input_start_col +
                                   year * 12 + month, value=getValue(year, month, "productive_stops"))

                    worksheet.cell(row=data_output_start_row - 1, column=data_output_start_col +
                                   year * 12 + month).value += str(company_budget_year+year)
                    worksheet.cell(row=data_output_start_row, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "total_consumption"))
                    
                    worksheet.cell(row=data_output_start_row + 1, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "utilization_factor_consistency_peakpower"))
                    worksheet.cell(row=data_output_start_row + 2, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "utilization_factor_consistency_offpeakpower"))
                    
                    worksheet.cell(row=data_output_start_row + 3, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "loadfactor_consistency_peakpower"))
                    worksheet.cell(row=data_output_start_row + 4, column=data_output_start_col + year *
                                   12 + month, value=getValue(year, month, "loadfactor_consistency_offpeakpower"))
                    
                    worksheet.cell(row=data_output_start_row + 5, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "uniqueload_factor_consistency"))
                    worksheet.cell(row=data_output_start_row + 6, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "modulation_factor_consistency"))
                    worksheet.cell(row=data_output_start_row + 7, column=data_output_start_col +
                                   year * 12 + month, value=getValue(year, month, "specific_consumption"))

        workbook_bytes = BytesIO()
        workbook.save(workbook_bytes)
        return workbook_bytes
