import csv

import pandas as pd
from django.http import HttpResponse
from io import StringIO, BytesIO
import pdfkit
from django.utils.encoding import smart_str
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from copy import copy
from base64 import b64encode


def get_leaves(item, key=None):
    if isinstance(item, dict):
        leaves = {}
        for i in item.keys():
            leaves.update(get_leaves(item[i], i))
        return leaves
    elif isinstance(item, list):
        leaves = {}
        for i in item:
            leaves.update(get_leaves(i, key))
        return leaves
    else:
        # put formatting here
        if item is None:
            item = ""
        if 'status' in key:
            if item == 'S':
                item = 'ATIVO'
            elif item == 'N':
                item = 'INATIVO'
        return {key: item}


def generic_data_csv_list(rest, ckeck_detail):
    '''

    :param rest: this field is an array with a dict. The rest it's the same as serializing
    :param ckeck_detail: array with values in dict that contains _detail.
                         If you pass a value, the method will bring in return
    :return: a new rest normalize and merge data

    '''
    def normalize_dict(**kwargs):
        for k in kwargs.keys():
            if type(kwargs[k]) is list:
                kwargs[k] = pd.io.json.json_normalize([kwargs], record_path=[k]).to_dict()
            if type(kwargs[k]) is dict:
                kwargs[k] = normalize_dict(
                    **dict([(smart_str(nk), nv) for nk, nv in kwargs[k].items()]))
        return kwargs

    def get_control(number, array, ckeck_detail, **kwargs):
        control = number
        key_normalize = array
        for keys in kwargs.keys():
            if type(kwargs[keys]) is dict:
                need_enter_again = True
                if (keys.__contains__('detail') and keys in ckeck_detail) or keys.find('detail') < 0:
                    try:
                        principal_key = list(kwargs[keys].keys())[0]
                        key_digit = list(kwargs[keys][principal_key])[0]
                        if '0' in kwargs[keys][principal_key] and key_digit.isdigit() and type(
                                kwargs[keys][principal_key]) is dict:
                            size = len(kwargs[keys][principal_key])
                            control = size if size > control else control
                            key_normalize.append(keys)
                            need_enter_again = False
                    except:
                        need_enter_again = True
                if need_enter_again:
                    control, key_normalize = get_control(control, key_normalize, ckeck_detail, **kwargs[keys])
        return control, key_normalize

    def generic_copy(key_normalize, **kwargs):
        core = kwargs.copy()
        for keys in kwargs.keys():
            if type(kwargs[keys]) is dict and keys not in key_normalize:
                core[keys] = generic_copy(key_normalize, **kwargs[keys])
            elif keys in key_normalize:
                del core[keys]
        return core

    rest_data = []
    for index in range(len(rest)):
        kwargs = normalize_dict(**rest[index])
        control = 0
        key_normalize = []
        control, key_normalize = get_control(control, key_normalize, ckeck_detail, **kwargs)
        core = generic_copy(key_normalize, **kwargs)
        if not control:
            rest_data.append(kwargs)
        else:
            for index in range(control):
                core_append = core.copy()
                for key, value_dict in kwargs.items():
                    if key in key_normalize:
                        core_append[key] = {}
                        for k, v in value_dict.items():
                            if str(index) in list(v.keys()):
                                core_append[key][k] = v[str(index)]
                            else:
                                core_append[key][k] = ""
                rest_data.append(core_append)
    return rest_data


def generic_pdf(mapping, header, rest_data, file_name, landscape=False):
    response = StringIO()
    html_file = StringIO()
    pdf_out = BytesIO()

    writer = csv.DictWriter(response, fieldnames=(mapping), delimiter=';')

    if bool(header):
        writer.writerow(header)
    else:  # pragma: no cover
        # optional from developer
        writer.writeheader()

    writer.writerows(get_leaves(entry) for entry in rest_data)

    response.seek(0)

    df = pd.read_csv(response, sep=';', keep_default_na=False)
    html_file = df.style.hide_index().set_table_styles(
       [{
           'selector': 'th',
           'props': [
               ('background-color', '#007E7A'),
               ('color', '#fff'),
               ('font-weight', 'bold')]
       },{
           'selector': '*',
           'props': [
               ('border', 'none')]
       }]).render()

    options = {
        'page-size': 'A4',  # to see sizes https://doc.qt.io/archives/qt-4.8/qprinter.html#PaperSize-enum
        'margin-top': '1cm',
        'margin-bottom': '1cm',
        'margin-left': '0.2cm',
        'margin-right': '0.2cm',
        'quiet': ''
    }

    if landscape:
        options["orientation"] = 'landscape'

    vale_logo_data = open("uploads/images/vale.png", "rb").read()
    vale_logo_encoded = b64encode(vale_logo_data).decode("utf-8")
    vale_logo_html = f"<img src='data:image/png;base64,{vale_logo_encoded}' width='150px'>"

    pdf_file = pdfkit.PDFKit('<meta charset="UTF-8">\n' + vale_logo_html + html_file, "string", options=options).to_pdf()

    response.close()

    response = HttpResponse(pdf_file)
    response['Content-Type'] = 'application/pdf'
    response['Content-disposition'] = 'attachment;filename={}.pdf'.format(file_name)
    return response


def generic_csv(mapping, header, rest_data, file_name):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename={}.csv'.format(file_name)
    # add capability especial characters
    response.write(u'\ufeff'.encode('utf8'))

    writer = csv.DictWriter(response, fieldnames=mapping, delimiter=';')

    if bool(header):
        writer.writerow(header)
    else:  # pragma: no cover
        # optional from developer
        writer.writeheader()

    writer.writerows(get_leaves(entry) for entry in rest_data)
    return response

def generic_xls(mapping, header, rest_data, file_name, styleColuns): #style type ==::: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(file_name)
    
    wb = Workbook()
    ws = wb.active
    ws.title='%s' %file_name

    vale_logo = Image("uploads/images/vale.png")
    vale_logo.anchor = 'A1'
    ws.add_image(vale_logo) 
    ws.row_dimensions[1].height = 48
    
    cont_colun=1
    cont_line=2
    for colun in mapping:
        ws.cell(row=cont_line, column=cont_colun).value=header[colun]
        ws.cell(row=cont_line, column=cont_colun).fill=PatternFill("solid", fgColor="007E7A")
        font = copy(ws.cell(row=cont_line, column=cont_colun).font)
        font.color="ffffff"
        font.b=True
        ws.cell(row=cont_line, column=cont_colun).font = font
        for data in rest_data:
            cont_line+=1
            valid_insert=True
            for valueStyle in styleColuns:
                if colun in valueStyle['fields']:
                    ws.cell(row=cont_line, column=cont_colun).value=data[colun]
                    if 'style' in valueStyle:
                        ws.cell(row=cont_line, column=cont_colun).style=valueStyle["style"]
                    if 'number_format' in valueStyle:
                        ws.cell(row=cont_line, column=cont_colun).number_format =valueStyle["number_format"]

                    valid_insert=False
                    break
            if valid_insert:
                ws.cell(row=cont_line, column=cont_colun).value=data[colun]

        cont_line=2
        cont_colun+=1

    wb.save(response)
    return response