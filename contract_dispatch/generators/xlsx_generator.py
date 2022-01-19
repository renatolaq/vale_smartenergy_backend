from io import BytesIO

from SmartEnergy.settings import BASE_DIR
from contract_dispatch.business.contract_dispatch_business import (
    ContractDispatchBusiness,
)
from rest_framework.exceptions import ValidationError
from openpyxl.styles import Font, PatternFill
import openpyxl


def generate_xlsx(request):
    contracts_by_month = ContractDispatchBusiness.retrieve_contracts_dict_list_by_month(
        request
    )
    result = generate(contracts_by_month)

    return result


def generate_xlsx_by_contract(request, contract_dispatch):
    contracts_by_month = ContractDispatchBusiness.retrieve_contracts_dict_list_by_contract_dispatch(
        request, contract_dispatch
    )
    result = generate(contracts_by_month)
    return result


def generate_xlsx_contracts_to_send(request, year, month, balance_id):
    contracts = ContractDispatchBusiness.retrieve_contracts_dict_list_contracts_to_send(
        request, year, month, balance_id
    )

    result = generate(contracts)
    return result


def generate_xlsx_contracts_by_ids(request, year, month, balance_id, ids):
    contracts = ContractDispatchBusiness.retrieve_contracts_dict_list_by_contract_cliq_ids(
        request, year, month, balance_id, ids
    )

    result = generate(contracts)
    return result


def generate(data):
    if data == []:
        raise ValidationError({"detail": f"No data to be exported."})

    workbook: openpyxl.Workbook = openpyxl.load_workbook(
        filename="contract_dispatch/templates/xlsx_template.xlsx"
    )
    logo_vale_folder = f"{BASE_DIR}/SmartEnergy/static/vale-logo.png"

    default_sheet = workbook.active
    for contracts_list in data:
        sheet_title = contracts_list["title"].replace("/", "-")
        logo_vale = openpyxl.drawing.image.Image(logo_vale_folder)

        current_sheet = workbook.copy_worksheet(default_sheet)
        current_sheet.title = sheet_title
        current_sheet.add_image(logo_vale)

        current_row = 2
        current_column = 1
        for field in contracts_list["fields"]:
            cell = current_sheet.cell(
                row=current_row, column=current_column, value=field["name"]
            )
            set_header_style(cell)
            current_column += 1
            pass

        current_row = 3
        current_column = 1
        for contract in contracts_list["contracts"]:
            for value in contract["values"]:
                cell = current_sheet.cell(
                    row=current_row, column=current_column, value=value
                )
                set_content_style(cell)
                current_column += 1
            current_row += 1
            current_column = 1

        adjust_column_cell(current_sheet)

    workbook.remove_sheet(default_sheet)

    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)

    workbook_bytes.seek(0)
    excelFile = workbook_bytes.read()

    return excelFile


def set_content_style(cell):
    font = Font(name="Calibri", size=11)

    cell.font = font


def set_header_style(cell):
    font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="327E7A", end_color="327E7A", fill_type="solid")

    cell.font = font
    cell.fill = fill


def adjust_column_cell(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value * 1.5
